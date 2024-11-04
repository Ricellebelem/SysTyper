import sqlite3
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, simpledialog
from tkinter import Tk, filedialog
from tkinter.filedialog import askopenfilename
from bs4 import BeautifulSoup
from tkinter import Toplevel
import ezdxf
from openpyxl import Workbook
import ctypes
import win32com.client as win32
from collections import defaultdict
from tkcalendar import DateEntry
import pandas as pd
from IPython.display import display
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime
import time
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException 
from selenium.webdriver.common.action_chains import ActionChains
import xlwings as xw
from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver.chrome.service import Service
import openpyxl
import pyautogui
import os
import shutil
from cryptography.fernet import Fernet
#==================================================================================================================================================================================================================================================================================================================================================
empresa = "Typer"
database_name = f'db_{empresa}.db'
#==================================================================================================================================================================================================================================================================================================================================================
def get_db_connection():
   conn = sqlite3.connect(database_name)
   conn.row_factory = sqlite3.Row
   return conn

def init_db():
   conn = get_db_connection()
   cursor = conn.cursor()
   cursor.execute('''CREATE TABLE IF NOT EXISTS processo (id INTEGER PRIMARY KEY, sob TEXT, pdl TEXT, solicitante TEXT, municipio TEXT, localidade TEXT, data_sol DATE, data_despacho DATE, servico TEXT, ceco TEXT, inspetor TEXT, tempo_enel INTEGER, estado_enel TEXT, motivo_reprovacao TEXT, obs_reprovacao TEXT, estudo TEXT, data_parceira DATE, data_gestor DATE, data_limite_12 DATE, data_limite_30 DATE, tempo_empresa INTEGER, estado_empresa TEXT, acao TEXT, responsavel TEXT, data_campo DATE, status_campo TEXT, tec_campo TEXT, motivo_susp_canc TEXT, obs_susp_canc TEXT, pendencia_enel TEXT, pendencia_parceira TEXT, data_desenho DATE, tec_desenho TEXT, data_orcamento DATE, tec_orcamento TEXT, data_revisao DATE, tec_revisao TEXT, data_documental DATE, tec_documental TEXT, data_traves_paralel DATE, tec_traves_paralel TEXT, data_entrega DATE, data_faturamento DATE, estado_02_03 TEXT, utm_coordinates TEXT, valuation TEXT, alim TEXT)''')
   conn.commit()
   conn.close()
init_db()

#==================================================================================================================================================================================================================================================================================================================================================
# Função para buscar dados do banco de dados com base no sob
def fetch_data():
    sob = entry_sob.get()
    conn = get_db_connection()    
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM processo WHERE sob=?", (sob,))
    row = cursor.fetchone()
    
    if row:
        entry_solicitante.delete(0, tk.END)
        entry_solicitante.insert(0, row['solicitante'])
        entry_municipio.delete(0, tk.END)
        entry_municipio.insert(0, row['municipio'])
    else:
        limpar_formularios()
        messagebox.showinfo("Atenção!", "Ordem não encontrada!")
    conn.close()

#==================================================================================================================================================================================================================================================================================================================================================
def excluir_tabela_processo():
    conn = sqlite3.connect(f"db_{empresa}.db")
    cursor = conn.cursor()
    cursor.execute('''DROP TABLE processo''')
#=====================================================trecho pertence a alteração de senha=========================================================================================================================================================================================================================================================
# Gera uma chave e salva em arquivo txt
def gerar_chave():
    chave = Fernet.generate_key()
    with open("chave.key", "wb") as chave_arquivo:
        chave_arquivo.write(chave)

# Carrega a chave de um arquivo
def carregar_chave():
    return open("chave.key", "rb").read()

# Criptografa os dados
def criptografar_dados(dados, chave):
    f = Fernet(chave)
    dados_encriptados = f.encrypt(dados.encode())
    return dados_encriptados

# Descriptografa os dados
def descriptografar_dados(dados_encriptados, chave):
    f = Fernet(chave)
    dados_decriptados = f.decrypt(dados_encriptados).decode()
    return dados_decriptados
#==================================================================================================================================================================================================================================================================================================================================================
def get_login_data(): # Função para obter dados de login (sem alterar)
    arquivo_login = "dados_login.txt"
    arquivo_chave = "chave.key"
    
    if not os.path.exists(arquivo_chave):
        gerar_chave()
    chave = carregar_chave()

    # Se o arquivo de login não existir, chama a função para alterar/cadastrar os dados
    if not os.path.exists(arquivo_login):
        print("Dados de login não encontrados. Por favor, cadastre.")
        return change_password_gom_gomnet()  # Chama para cadastrar novos dados
    
    # Se o arquivo existe, carrega e descriptografa
    with open(arquivo_login, "rb") as arquivo:
        login_criptografado = arquivo.readline().strip()
        senha_criptografada = arquivo.readline().strip()

    login = descriptografar_dados(login_criptografado, chave)
    senha = descriptografar_dados(senha_criptografada, chave)
    return login, senha
#==================================================================================================================================================================================================================================================================================================================================================
def change_password_gom_gomnet(): # Função para alterar ou cadastrar dados de login - Comando_1
    arquivo_login = "dados_login.txt"
    arquivo_chave = "chave.key"

    if not os.path.exists(arquivo_chave):
      gerar_chave()
    chave = carregar_chave()
    
    if os.path.exists(arquivo_login):
      resposta = messagebox.askyesno("Typer", "Deseja alterar os dados de login?")
      if not resposta:
        return get_login_data()  # Retorna dados atuais sem alteração

    # Solicita novo login e senha
    login = simpledialog.askstring("Typer", "Digite o novo BR:")
    senha = simpledialog.askstring("Typer", "Digite a nova senha:")

    # Criptografa e grava os novos dados
    login_criptografado = criptografar_dados(login, chave)
    senha_criptografada = criptografar_dados(senha, chave)

    with open(arquivo_login, "wb") as arquivo:
        arquivo.write(login_criptografado + b"\n")
        arquivo.write(senha_criptografada + b"\n")
    
    messagebox.showinfo("Sucesso!", "Dados de login gravados com sucesso!")
    return login, senha 
#==================================================================================================================================================================================================================================================================================================================================================
def exportar_tabela_processo():
    conn = sqlite3.connect(f"db_{empresa}.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, sob, pdl, solicitante, municipio, localidade, strftime('%d/%m/%Y', data_sol) AS data_sol, strftime('%d/%m/%Y', data_despacho) AS data_despacho, servico, ceco, inspetor, tempo_enel, estado_enel, motivo_reprovacao, obs_reprovacao, estudo, strftime('%d/%m/%Y', data_parceira) AS data_parceira, strftime('%d/%m/%Y', data_gestor) AS data_gestor, strftime('%d/%m/%Y', data_limite_12) AS data_limite_12, strftime('%d/%m/%Y', data_limite_30) AS data_limite_30, tempo_empresa, estado_empresa, acao, responsavel, strftime('%d/%m/%Y', data_campo) AS data_campo, status_campo, tec_campo, motivo_susp_canc, obs_susp_canc, pendencia_enel, pendencia_parceira, strftime('%d/%m/%Y', data_desenho) AS data_desenho, tec_desenho, strftime('%d/%m/%Y', data_orcamento) AS data_orcamento, tec_orcamento, strftime('%d/%m/%Y', data_revisao) AS data_revisao, tec_revisao, strftime('%d/%m/%Y', data_documental) AS data_documental, tec_documental, strftime('%d/%m/%Y', data_traves_paralel) AS data_traves_paralel, tec_traves_paralel, strftime('%d/%m/%Y', data_entrega) AS data_entrega, strftime('%d/%m/%Y', data_faturamento) AS data_faturamento, estado_02_03, utm_coordinates, valuation, alim FROM processo")
    registros = cursor.fetchall()
    colunas = [description[0] for description in cursor.description]
    # Criar um DataFrame a partir dos registros e das colunas
    df = pd.DataFrame(registros, columns=colunas)
    tabela_processo = f"{empresa}_tabela_processo.xlsx"
    df.to_excel(tabela_processo,index=False)
    conn.close()
#==================================================================================================================================================================================================================================================================================================================================================
def limpar_formularios():
   entry_solicitante.delete(0, tk.END)
   entry_municipio.delete(0, tk.END)
#==================================================================================================================================================================================================================================================================================================================================================
def deletar_ordem_tkinter():
    sob = entry_sob.get()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM processo WHERE sob=?", (sob,))
    row = cursor.fetchone()
    if row:
      cursor.execute("DELETE FROM processo WHERE sob = ?", (sob,))
      cursor.execute("SELECT * FROM processo WHERE sob=?", (sob,))
      row = cursor.fetchone()
      if row:
        messagebox.showinfo("Atenção!", "Ordem não excluída!")
      else:
        limpar_formularios()
        messagebox.showinfo("Sucesso!", "Ordem excluída com sucesso!")
    else:
      messagebox.showinfo("Atenção!", "Ordem não encontrada!")
    
    conn.commit()
    conn.close()
#==================================================================================================================================================================================================================================================================================================================================================    
def import_relat_vistoria():
    # Solicitar ao usuário para escolher uma planilha
    planilha_vistoria = askopenfilename(title="Selecione a planilha", filetypes=[("Excel files", "*.xlsx *.xls")])
    
    if not planilha_vistoria:
        print("Nenhuma planilha selecionada.")
        return

    # Ler a planilha
    df = pd.read_excel(planilha_vistoria)
    df_duplicadas = df[df.duplicated(subset='SLNS_CODIGO', keep=False)]
    
    if df_duplicadas.empty:
      # Conectar ao banco de dados (ou criar um novo)
      conn = sqlite3.connect(f"db_{empresa}.db")
      cursor = conn.cursor()

      # Verificar quais SOBs já existem no banco de dados
      sobs_existentes = []
      for sob in df['SLNS_CODIGO']:
        sob = str(sob).zfill(10)
        cursor.execute("SELECT sob FROM processo WHERE sob = ?", (sob,))
        if cursor.fetchone():
          sobs_existentes.append(sob)

      if sobs_existentes:
        # Salvar SOBs existentes em um arquivo Excel
        df_existentes = df[df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(sobs_existentes)]
        df_existentes.to_excel("sobs_existentes.xlsx", index=False)
        
        # Informar ao usuário sobre os SOBs existentes
        messagebox.showinfo("Atenção!", "Alguns sobs já existem no banco de dados. verificar arquivo 'sobs_existentes.xlsx'.")
       
        # Perguntar ao usuário se deseja prosseguir com a gravação dos novos dados
        if messagebox.askyesno("Continuar?", "Deseja prosseguir com a gravação apenas dos dados cujos SOBs ainda não existem no banco de dados?"):
          # Filtrar dados que não existem no banco de dados
          df_novos = df[~df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(sobs_existentes)]
          # Inserir novos dados no banco de dados
          if df_novos.empty:
            messagebox.showinfo("Atenção!", "Update não realizado. Não existem dados novos para atualizar.")
          else:
            inserir_relat_vist_no_banco(df_novos, cursor)
            conn.commit()            
            messagebox.showinfo("Sucesso!", "Novos dados importados com sucesso!")
        else:
          messagebox.showinfo("Atenção!", "Importação de dados cancelada!")
      else:
        inserir_relat_vist_no_banco(df, cursor)
        conn.commit()
        messagebox.showinfo("Sucesso!", "Todos os dados importados com sucesso!")
        
      # Mostrar todos os dados do banco de dados
      cursor.execute("SELECT * FROM processo")
      registros = cursor.fetchall()
      colunas = [description[0] for description in cursor.description]
      df_db = pd.DataFrame(registros, columns=colunas)
      #print(df_db)  # Exibe os dados no console (ou use display(df_db) no Jupyter Notebook)
      conn.close()
    else:
        df_duplicadas.to_excel("sobs_duplicadas.xlsx", index=False)
        messagebox.showinfo("Atenção!", "Update não realizado. Existem SOBs duplicadas na planilha. Verifique o arquivo 'sobs_duplicadas.xlsx'.")

def inserir_relat_vist_no_banco(df, cursor):
    for _, row in df.iterrows():
        sob = str(row['SLNS_CODIGO'])
        pdl = str(row['NUM_ORDEM'])
        
        # Ajuste de zeros à esquerda
        sob = sob.zfill(10) if sob.isdigit() else sob.rjust(10, '0')
        pdl = pdl.zfill(10) if pdl.isdigit() else pdl.rjust(10, '0')
        
        cursor.execute('''INSERT INTO processo (
            sob, pdl, solicitante, municipio, localidade, data_sol, data_despacho, servico, ceco, inspetor, tempo_enel, estado_enel
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
            sob, 
            pdl,
            row['NOME_CLIENTE'], 
            row['MUNICIPIO'], 
            row['LOCALIDADE'],
            str(row['DATA_SOLIC'])[:10],
            str(row['SODE_FECHA_DESPACHO'])[:10],                    
            row['TIPO_SERVICO'], 
            row['CODIGO'], 
            row['INSPETOR_NOM'], 
            row['DURACAO'],
            '2'))
#==================================================================================================================================================================================================================================================================================================================================================    
def import_relat_02_e_03():
    # Solicitar ao usuário para escolher uma planilha
    planilha_02_e_03 = askopenfilename(title="Selecione a planilha", filetypes=[("Excel files", "*.xlsx *.xls")])
    
    if not planilha_02_e_03:
        print("Nenhuma planilha selecionada.")
        return

    # Ler a planilha
    df = pd.read_excel(planilha_02_e_03)
    df_duplicadas = df[df.duplicated(subset='SLNS_CODIGO', keep=False)]
    
    if df_duplicadas.empty:
        # Conectar ao banco de dados (ou criar um novo)
        conn = sqlite3.connect(f"db_{empresa}.db")
        cursor = conn.cursor()
    
        # Verificar quais SOBs já existem no banco de dados
        sobs_existentes = []
        for sob in df['SLNS_CODIGO']:
            sob = str(sob).zfill(10)
            cursor.execute("SELECT sob FROM processo WHERE sob = ?", (sob,))
            if cursor.fetchone():
                sobs_existentes.append(sob)

        if sobs_existentes:
            # Salvar SOBs existentes em um arquivo Excel
            df_existentes = df[df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(sobs_existentes)]
            df_existentes.to_excel("sobs_existentes.xlsx", index=False)
        
            # Informar ao usuário sobre os SOBs existentes
            messagebox.showinfo("Atenção!", "Alguns SOBs já existem no banco de dados. Verificar arquivo 'sobs_existentes.xlsx'.")
        
            # Perguntar ao usuário se deseja prosseguir com a gravação dos novos dados
            if messagebox.askyesno("Continuar?", "Deseja prosseguir com a gravação apenas dos dados cujos SOBs ainda não existem no banco de dados?"):
                # Filtrar dados que não existem no banco de dados
                df_novos = df[~df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(sobs_existentes)]
                if df_novos.empty:
                   messagebox.showinfo("Atenção!", "Update não realizado. Não existem dados novos para atualizar.")
                else:
                  inserir_relat_02_e_03_no_banco(df_novos, cursor)
                  conn.commit()
                  messagebox.showinfo("Sucesso!", "Novos dados importados com sucesso!")
            else:
                messagebox.showinfo("Atenção!", "Importação de dados cancelada!")
        else:
          inserir_relat_02_e_03_no_banco(df, cursor)
          conn.commit()
          messagebox.showinfo("Sucesso!", "Todos os dados importados com sucesso!")
          
        cursor.execute("SELECT * FROM processo")
        registros = cursor.fetchall()
        colunas = [description[0] for description in cursor.description]
        # Criar um DataFrame a partir dos registros e das colunas
        df_db = pd.DataFrame(registros, columns=colunas)
        # Exibir o DataFrame
        print(df_db)
        conn.close()
    else:
        df_duplicadas.to_excel("sobs_duplicadas.xlsx", index=False)
        messagebox.showinfo("Atenção!", "Update não realizado. Existem sobs duplicadas na planilha. Verificar arquivo sobs_duplicadas.xlsx.")

def inserir_relat_02_e_03_no_banco(df, cursor):
  for _, row in df.iterrows():
    sob = str(row['SLNS_CODIGO'])
    pdl = str(row['SOB'])
    # Ajuste de zeros à esquerda
    sob = sob.zfill(10) if sob.isdigit() else sob.rjust(10, '0')
    pdl = pdl.zfill(10) if pdl.isdigit() else pdl.rjust(10, '0')
    cursor.execute('''INSERT INTO processo (sob, pdl, solicitante, municipio, data_sol, data_despacho, servico, ceco, inspetor, tempo_enel, estado_enel, motivo_reprovacao, obs_reprovacao, estudo, data_parceira, data_gestor) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) ''', (
      sob,
      pdl,
      row['NOMECLIENTE'],
      row['MUNICIPIO'],
      str(row['DATASOLIC']) [:10],
      str(row['DATA_DESPACHO']) [:10],
      row['TIPO_SERVICO'],
      row['DEPTO'],
      row['INSP_NOME'],
      row['DURACAO'],
      row['ESTADOSOLIC'],
      row['MOTIVO_REPROVACAO'],
      row['OBS_REPROVACAO'],
      row['PROY_CODIGO'],
      str(row['PROY_FECHA_RETORNO_EMPREITEIRA']) [:10],
      str(row['PROY_FECHA_FIN_ACEPT_RECHAZO']) [:10]))

#==================================================================================================================================================================================================================================================================================================================================================    
def verificar_localidade():

  # Conectar ao banco de dados (ou criar um novo)
  conn = sqlite3.connect(f"db_{empresa}.db")
  cursor = conn.cursor()

  # Verificar quais "sob" têm o campo "localidade" vazio
  cursor.execute("SELECT sob FROM processo WHERE localidade IS NULL OR localidade = ''")
  result = cursor.fetchall()

  # Se houver resultados, salvar em um arquivo Excel e mostrar uma mensagem
  if result:
    # Converter o resultado para um DataFrame do pandas
    arquivo_vazio = f"{empresa}_ordens_sem_localidade.xlsx"
    df = pd.DataFrame(result, columns=['SLNS_CODIGO'])
    df ['LOCALIDADE'] = ""
    # Salvar o DataFrame em um arquivo Excel
    df.to_excel(arquivo_vazio,index=False)
           
    # Mostrar uma caixa de mensagem informando que existem ordens sem a informação da localidade
    messagebox.showinfo("Atenção!", "Existem ordens sem a informação da localidade. Verificar arquivo ordens_sem_localidade.xlsx")
  else:
    # Mostrar uma caixa de mensagem informando que todas as ordens têm a informação da localidade
    messagebox.showinfo("Atenção!", "Todas as ordens possuem a informação da localidade!")
    
  # Fechar a conexão com o banco de dados
  conn.close()

#==================================================================================================================================================================================================================================================================================================================================================
def atualizar_localidades():
    # Solicitar ao usuário para escolher uma planilha
    planilha_localidades = askopenfilename(title="Selecione uma planilha com as colunas SLNS_CODIGO e LOCALIDADE", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not planilha_localidades:
        print("Nenhuma planilha selecionada.")
        return

    # Ler a planilha
    df = pd.read_excel(planilha_localidades, usecols=["SLNS_CODIGO", "LOCALIDADE"])
    df_duplicadas = df[df.duplicated(subset='SLNS_CODIGO', keep=False)]
    
    if df_duplicadas.empty:
      # Conectar ao banco de dados (ou criar um novo)
      conn = sqlite3.connect(f"db_{empresa}.db")
      cursor = conn.cursor()

      #verificar se lovalidades vazio
      localidades_existentes = []
      for sob in df['SLNS_CODIGO']:
          sob = str(sob).zfill(10)
          cursor.execute("SELECT sob FROM processo WHERE sob = ? AND localidade IS NOT NULL", (sob,))
          if cursor.fetchone():
            localidades_existentes.append(sob)

      if localidades_existentes:
         # Salvar SOBs das localidades existentes em um arquivo Excel
            df_existentes = df[df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(localidades_existentes)]
            df_existentes.to_excel("localidades_existentes.xlsx", index=False)
        
            # Informar ao usuário sobre os SOBs existentes
            messagebox.showinfo("Atenção!", "Algumas localidades já existem no banco de dados. Verificar arquivo 'localidades_existentes.xlsx'.")
            
            # Perguntar ao usuário se deseja prosseguir com a gravação dos novos dados
            if messagebox.askyesno("Continuar?", "Deseja prosseguir com a gravação apenas dos dados cujas localidades ainda não existem no banco de dados?"):
                # Filtrar dados que não existem no banco de dados
                df_loc_novos = df[~df['SLNS_CODIGO'].astype(str).str.zfill(10).isin(localidades_existentes)]
                if df_loc_novos.empty:
                   messagebox.showinfo("Atenção!", "Update não realizado. Não existem dados novos para atualizar.")
                else:
                  inserir_localidades_no_banco(df_loc_novos, cursor)
                  conn.commit()
                  messagebox.showinfo("Sucesso!", "Novos dados importados com sucesso!")
            else:
                messagebox.showinfo("Atenção!", "Importação de dados cancelada!")
      else:
        inserir_localidades_no_banco(df, cursor)
        conn.commit()
        messagebox.showinfo("Sucesso!", "Todos os dados importados com sucesso!")
      
        
      cursor.execute("SELECT sob, localodade FROM processo WHERE sob = ?", (sob,))
      registros = cursor.fetchall()
      colunas = [description[0] for description in cursor.description]
      # Criar um DataFrame a partir dos registros e das colunas
      df_db = pd.DataFrame(registros, columns=colunas)
      # Exibir o DataFrame
      print(df_db)
      conn.close()
    
    else:
      df_duplicadas.to_excel("localidades_duplicadas.xlsx", index=False)
      messagebox.showinfo("Atenção!", "Update não realizado. Existem localidades duplicadas na planilha. Verificar arquivo localidades_duplicadas.xlsx.")



def inserir_localidades_no_banco(df, cursor):
   print('ob')
   '''
   # Atualizar as localidades no banco de dados
      for index, row in df.iterrows():
        sob = row["SLNS_CODIGO"]
        localidade = row["LOCALIDADE"]

        # Verificar se o campo "localidade" está vazio
        cursor.execute("SELECT localidade FROM processo WHERE sob = ?", (sob,))
        result = cursor.fetchone()

        if result and not result[0]:  # Se o campo "localidade" estiver vazio
            cursor.execute("UPDATE processo SET localidade = ? WHERE sob = ?", (localidade, sob))
            print(f"Atualizado: {sob} -> {localidade}")
    
    
      # Confirmar as mudanças
      conn.commit()
      conn.close()
         
      messagebox.showinfo("Sucesso!", "Atualização concluída!") 
    '''
#==========================================================================================================================
def atualizar_localidades_gom():
    # Conectar ao banco de dados
    conn = sqlite3.connect(f"db_{empresa}.db")
    cursor = conn.cursor()

    # Verificar quais "sob" têm o campo "localidade" vazio
    cursor.execute("SELECT sob FROM processo WHERE localidade IS NULL OR localidade = ''")
    result = cursor.fetchall()

    if result:
        arquivo_vazio = f"{empresa}_ordens_sem_localidade.xlsx"
        df = pd.DataFrame(result, columns=['SLNS_CODIGO'])
        df ['LOCALIDADE'] = ""
        df.to_excel(arquivo_vazio,index=False)
        
        url_inicial = 'http://gom-ce.enelint.global/loginweb/login.asp' #Tela de login
        url_ordem = 'http://gom-ce.enelint.global/AyudaActiva/AA_Consulta_Solicitud_NNSS.asp' #Tela para informar a sob
        login, senha = get_login_data()

        print(login)
        print(senha)
       
        driver = webdriver.Chrome()
        driver.get(url_inicial)
        elem = driver.find_element(By.NAME,'Usuario')
        elem.clear()
        elem.send_keys(login)
        elem = driver.find_element(By.NAME,'Clave')
        elem.clear()
        elem.send_keys(senha)
        elem.send_keys(Keys.RETURN)
        time.sleep(1)
        driver.get(url_ordem)
        print('pagina da sob')

        df_arquivo = pd.read_excel(arquivo_vazio, usecols=['SLNS_CODIGO', 'LOCALIDADE'])
        df_arquivo['LOCALIDADE'] = df_arquivo['LOCALIDADE'].astype(str)  # Converte a coluna para string


        sobs_inexistentes_na_gom = []
        for index, row in df_arquivo.iterrows():
            sob = row['SLNS_CODIGO']
            # Garantir que `sob` seja uma string com 10 caracteres
            if sob.isdigit():  # Caso `sob` seja apenas números
              sob = sob.zfill(10)  # Preenche com zeros à esquerda até ter 10 dígitos
            else:
              sob = sob.rjust(10, '0')  # Para valores alfanuméricos, preenche à esquerda até 10 caracteres
            
            try:
                elem = WebDriverWait(driver, 10).until (EC.presence_of_element_located((By.NAME, 'txtSlnsCodigo')))
                elem.clear()
                elem.send_keys(sob)
            except TimeoutException:
                print('Campo SOB não encontrado')

            elem = driver.find_element(By.NAME, 'txtPfisUser')
            elem.clear()
            elem.send_keys(Keys.RETURN)
                                 
            try:
                elem_loc = WebDriverWait(driver, 10).until (EC.presence_of_element_located((By.XPATH, '//*[@id="Body"]/form/table[2]/tbody/tr[1]/td[3]/table[7]/tbody/tr[2]/td[6]/a')))
                localidade = elem_loc.text
                df_arquivo.at[index, 'LOCALIDADE'] = str(localidade)  # Certifica-se de que o valor também é string
            except TimeoutException:
                sobs_inexistentes_na_gom.append(index)
                print(f'Localidade não encontrada para o SOB {sob}')
        if sobs_inexistentes_na_gom:
            df_sobs_inexistentes_na_gom =  df_arquivo.loc[sobs_inexistentes_na_gom, ['SLNS_CODIGO']].copy()
            df_sobs_inexistentes_na_gom['Erro'] = 'sob inexistentes na gom'
            df_sobs_inexistentes_na_gom.to_excel('sobs_inexistentes_na_gom.xlsx',index=False)
            # Informar ao usuário sobre os SOBs existentes
            messagebox.showinfo("Atenção!", "Algumas sobs não foram encontradas na gom. Verificar arquivo 'sobs_inexistentes_na_gom.xlsx'.")
                              
        # Salvar o DataFrame atualizado
        df_arquivo.to_excel(arquivo_vazio, index=False)
        print(f'Arquivo Excel atualizado em {arquivo_vazio}')
        driver.quit()

        df = pd.read_excel(arquivo_vazio, usecols=["SLNS_CODIGO", "LOCALIDADE"])

        # Conectar ao banco de dados (ou criar um novo)
        conn = sqlite3.connect(f"db_{empresa}.db")
        cursor = conn.cursor()

        # Atualizar as localidades no banco de dados
        for index, row in df.iterrows():
            sob = row["SLNS_CODIGO"]
            localidade = row["LOCALIDADE"]

            # Verificar se o campo "localidade" está vazio
            cursor.execute("SELECT localidade FROM processo WHERE sob = ?", (sob,))
            result = cursor.fetchone()

            if result and not result[0]:  # Se o campo "localidade" estiver vazio
                cursor.execute("UPDATE processo SET localidade = ? WHERE sob = ?", (localidade, sob))
                print(f"Atualizado: {sob} -> {localidade}")
                conn.commit()
                       
    else:
        # Mostrar uma caixa de mensagem informando que todas as ordens têm a informação da localidade
        messagebox.showinfo("Atenção!", "Todas as ordens possuem a informação da localidade!")
            
    conn.close()
    messagebox.showinfo("Sucesso!", "Atualização concluída!")
#==========================================================================================================================
def download_reports_02_e_03(): #Baixar retatórios da GomNet do estado 2 e 3 da gerência sul - Comando_2
  # Verificar se o arquivo de dados de login existe
  #change_password_gom_gomnet()
  
  """
  arquivo_login = "dados_login.txt"

  if not os.path.exists(arquivo_login):
    # Caso o arquivo não exista, solicitar o login e a senha via caixa de diálogo
    login = simpledialog.askstring("Typer", "Digite o BR:")
    senha = simpledialog.askstring("Typer", "Digite a senha:")

    # Gravar os dados de login no arquivo
    with open(arquivo_login, "w") as arquivo:
      arquivo.write(f"{login}\n")
      arquivo.write(f"{senha}\n")
    messagebox.showinfo("Sucesso!", "Dados de login gravados com sucesso!")
  
  else:
    # Ler os dados de login do arquivo
    with open(arquivo_login, "r") as arquivo:
      login = arquivo.readline().strip()
      senha = arquivo.readline().strip()
  """
  #messagebox.showinfo("Atenção!", "Antes de prosseguir, certifique-se de que não haja relatórios pré-existentes no diretório (C:\\Typer\\RELATORIOS)!")
  
  # Inicialização do WebDriver do Microsoft Edge
  options = EdgeOptions()
  options.use_chromium = True
  driver = Edge(options=options)
  urllogin = 'http://gomnet-ce.enelint.global/'
  urlacomp_proj = 'http://gomnet-ce.enelint.global/EstudoTecnico.aspx'
  login, senha = get_login_data()
  
  #chrome_driver_path = 'C:\\Typer\\chromedriver.exe'
  #service = Service(chrome_driver_path)
  #driver = webdriver.Chrome(service=service)
  
  # Navega para a página de login
  driver.get(urllogin)
  # Localiza e preenche o campo de login
  campo_login = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'txtBoxLogin')))
  campo_login.send_keys(login)
  # Localiza e preenche o campo de senha
  campo_senha = driver.find_element(By.ID, 'txtBoxSenha') 
  campo_senha.send_keys(senha)
  # Localiza e clica no botão de login
  botao_login = driver.find_element(By.XPATH, '//*[@id="ImageButton_Login"]')  # Substitua pelo XPath correto
  botao_login.click()

  wait = WebDriverWait(driver, 10)
  elemento_alvo = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/form/table/tbody/tr[4]/td/div/div[1]/div/table/tbody/tr[5]/td/table/tbody/tr/td/input')))  # Substitua pelo ID correto

  def baixar():    
    driver.get (urlacomp_proj)
    campo_ceco = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_TextBox_Ceco') 
    campo_ceco.send_keys(ceco)
    campo_estado = driver.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_DropDownList_EstadoSolicitacao') 
    campo_estado.send_keys('solicitado estudo tecnico')
    botao_exportar = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_ImageButton_Excel')  # Substitua pelo XPath correto
    botao_exportar.click()
    time.sleep(2)
    
    pasta_origem = r'C:\\Downloads'
    pasta_destino = r'C:\\Typer\\RELATORIOS'
    # Obtém uma lista de todos os arquivos na pasta de origem
    arquivos = os.listdir(pasta_origem)
    # Filtra apenas os arquivos (excluindo pastas)
    arquivos = [arquivo for arquivo in arquivos if os.path.isfile(os.path.join(pasta_origem, arquivo))]
    # Classifica os arquivos por data de modificação (do mais recente para o mais antigo)
    arquivos.sort(key=lambda x: os.path.getmtime(os.path.join(pasta_origem, x)), reverse=True)
    
    # Verifica se existem arquivos na pasta
    if arquivos:
      # Obtém o caminho completo do último arquivo baixado
      ultimo_arquivo = os.path.join(pasta_origem, arquivos[0])
      # Move o último arquivo baixado para a pasta de destino
      shutil.move(ultimo_arquivo, pasta_destino)
      time.sleep(1)
      caminho_arquivo = 'C:\\Typer\\RELATORIOS\\Relatorio_EstudoTecnico.xls'
      novo_nome_arquivo = 'C:\\Typer\\RELATORIOS\\' + str(ceco) + ' 02.xls'
      # Renomear o arquivo
      os.rename(caminho_arquivo, novo_nome_arquivo)
    else:
      print("Nenhum arquivo encontrado na pasta de origem.")

    driver.get (urlacomp_proj)
    campo_ceco = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_TextBox_Ceco') 
    campo_ceco.send_keys(ceco)
    campo_estado = driver.find_element(By.CSS_SELECTOR, '#ctl00_ContentPlaceHolder1_DropDownList_EstadoSolicitacao') 
    campo_estado.send_keys('iniciado estudo tecnico')
    botao_exportar = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_ImageButton_Excel')  # Substitua pelo XPath correto
    botao_exportar.click()
    time.sleep(2)

    pasta_origem = r'C:\\Downloads'
    pasta_destino = r'C:\\Typer\\RELATORIOS'
    # Obtém uma lista de todos os arquivos na pasta de origem
    arquivos = os.listdir(pasta_origem)
    # Filtra apenas os arquivos (excluindo pastas)
    arquivos = [arquivo for arquivo in arquivos if os.path.isfile(os.path.join(pasta_origem, arquivo))]
    # Classifica os arquivos por data de modificação (do mais recente para o mais antigo)
    arquivos.sort(key=lambda x: os.path.getmtime(os.path.join(pasta_origem, x)), reverse=True)
 
    # Verifica se existem arquivos na pasta
    if arquivos:
      # Obtém o caminho completo do último arquivo baixado
      ultimo_arquivo = os.path.join(pasta_origem, arquivos[0])
      # Move o último arquivo baixado para a pasta de destino
      shutil.move(ultimo_arquivo, pasta_destino)
      time.sleep(1)
      caminho_arquivo = 'C:\\Typer\\RELATORIOS\\Relatorio_EstudoTecnico.xls'
      novo_nome_arquivo = 'C:\\Typer\\RELATORIOS\\' + str(ceco) + ' 03.xls'
      # Renomear o arquivo
      os.rename(caminho_arquivo, novo_nome_arquivo)

    else:
      print("Nenhum arquivo encontrado na pasta de origem.")
      time.sleep(2)

  ceco = "8531"
  baixar()

  ceco = "8521"
  baixar()

  ceco = "8511"
  baixar()

  time.sleep (2)

  driver.quit()

  # Sera necessario instalar o BeautifulSoup, pandas e xlsxwriter
  # pip install bs4 pandas xlsxwriter

  # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8511 02.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8511 02.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
      lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()

    # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8511 03.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8511 03.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
      lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()

  # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8521 02.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8521 02.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
     lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()
    # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8521 03.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8521 03.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
      lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()

    # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8531 02.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8531 02.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
      lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()

    # Preencher de acordo com o local do arquivo
  arquivo_entrada = "C:\\Typer\\RELATORIOS\\8531 03.xls"
  arquivo_saida = "C:\\Typer\\RELATORIOS\\8531 03.xlsx"

  with open(arquivo_entrada) as html_file:
    # ler o arquivo como pagina html
    soup = BeautifulSoup(html_file, 'html.parser')
    lista_linhas=[]
    lista_colunas=[]
 
    # Buscar tabela pelo ID dela
    table = soup.find(id='GridView1_ctl02_GridView_Excel')
 
    # Extrair os cabeçalhos da tabela e inserir em uma lista
    for th in table.find_all('th'):
      lista_colunas.append(th.text)
    lista_linhas.append(lista_colunas)


    # Extrair as colunas e linhas da tabela e inserir em uma lista
    for tr in table.find_all('tr'):
      lista_colunas=[]
      tds = tr.find_all('td')
      # Verificar se a tabela tem pelo menos uma coluna e se a primeira coluna não é vazia
      if len(tds) == 0 or tds[0].text == '':
        continue

      for td in tr.find_all('td'):
        valor = td.text
        # opcional: convertendo o td para inteiro para corrigir o erro de formato no excel
        try:
          valor = int(valor)
        except ValueError:
          pass
        # fim opcional
        lista_colunas.append(valor)
      lista_linhas.append(lista_colunas)
      
    df=pd.DataFrame(lista_linhas)

    # Gera a planilha de saida
    writer = pd.ExcelWriter(arquivo_saida, engine="xlsxwriter")
    df.to_excel(writer, sheet_name="Planilha1", header=False, index=False)

    workbook = writer.book
    worksheet = writer.sheets["Planilha1"]

    (max_row, max_col) = df.shape

    # Cria uma lista com os cabeçalhos das colunas
    colunas = [{"header": str(column)} for column in lista_linhas[0]]
    print(colunas)

    # Adiciona a tabela na planilha usando as colunas
    worksheet.add_table(0, 0, max_row - 1, max_col - 1, {"columns": colunas})

    # Ajusta a largura das colunas
    worksheet.set_column(0, max_col - 1, 12)

    #Fechar o arquivo de saida
    writer.close()

  time.sleep(1)

  # Especifique o caminho completo para o arquivo que você deseja excluir
  caminho_8511_02 = "C:\\Typer\\RELATORIOS\\8511 02.xls"
  caminho_8511_03 = "C:\\Typer\\RELATORIOS\\8511 03.xls"
  caminho_8521_02 = "C:\\Typer\\RELATORIOS\\8521 02.xls"
  caminho_8521_03 = "C:\\Typer\\RELATORIOS\\8521 03.xls"
  caminho_8531_02 = "C:\\Typer\\RELATORIOS\\8531 02.xls"
  caminho_8531_03 = "C:\\Typer\\RELATORIOS\\8531 03.xls"

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8511_02):
    os.remove(caminho_8511_02)
  else:
    print(f'O arquivo {caminho_8511_02} não existe.')

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8511_03):
    os.remove(caminho_8511_03)
  else:
    print(f'O arquivo {caminho_8511_03} não existe.')

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8521_02):
    os.remove(caminho_8521_02)
  else:
    print(f'O arquivo {caminho_8521_02} não existe.')

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8521_03):
    os.remove(caminho_8521_03)
  else:
    print(f'O arquivo {caminho_8521_03} não existe.')

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8531_02):
    os.remove(caminho_8531_02)
  else:
    print(f'O arquivo {caminho_8531_02} não existe.')

  # Verifique se o arquivo existe antes de tentar excluí-lo
  if os.path.exists(caminho_8531_03):
    os.remove(caminho_8531_03)
  else:
    print(f'O arquivo {caminho_8531_03} não existe.')

    
  #Unificando os arquivos em um único relatório

  def unir_arquivos_xlsx(diretorio, arquivo_saida):
    # Listar todos os arquivos xlsx no diretório
    arquivos_xlsx = [arquivo for arquivo in os.listdir(diretorio) if arquivo.endswith('.xlsx')]

    # Verificar se há pelo menos dois arquivos para unir
    if len(arquivos_xlsx) < 2:
      print("É necessário pelo menos dois arquivos para unir.")
      return

    # Criar um DataFrame vazio
    df_final = pd.DataFrame()

    # Iterar sobre os arquivos e concatenar
    for arquivo in arquivos_xlsx:
        caminho_arquivo = os.path.join(diretorio, arquivo)
        df_temporario = pd.read_excel(caminho_arquivo)
        df_final = pd.concat([df_final, df_temporario], ignore_index=True)

        # Salvar o DataFrame final no arquivo de saída
        df_final.to_excel(arquivo_saida, index=False)
        print(f'Arquivos unidos com sucesso. Saída salva em: {arquivo_saida}')

  # Exemplo de uso
  unir_arquivos_xlsx('C:\\Typer\\RELATORIOS', 'C:\\Typer\\RELATORIOS\\ordens.xlsx')
    
  messagebox.showinfo("Sucesso!", "Relatórios baixados com sucesso!")
#==================================================================================================================================================================================================================================================================================================================================================
def extrair_do_dxf(): # Comando_3
    # Inicializa a janela do Tkinter
    root = Tk()
    root.withdraw()  # Esconde a janela principal
    
    # Abre a janela de seleção de arquivos
    dwg_file = askopenfilename(filetypes=[("Arquivos DWG", "*.dwg")], title="Selecione o arquivo DWG")
    if not dwg_file:
        return
    
    excel_dados = "C:\\Typer\\DADOS.xlsx"
    
    try:
        gstar_app = win32.GetObject(None, "GstarCAD.Application")
    except:
        gstar_app = win32.Dispatch("GstarCAD.Application")
    
    gstar_doc = gstar_app.Documents.Open(dwg_file)
    
    # Verifica se o arquivo DADOS.xlsx existe
    if os.path.exists(excel_dados):
        wb = openpyxl.load_workbook(excel_dados)
    else:
        wb = Workbook()
        wb.create_sheet(title="DADOS")
        wb.create_sheet(title="GERAL")
        wb.remove(wb["Sheet"])  # Remove a aba padrão criada automaticamente
    
    ws_dados = wb["DADOS"]
    ws_geral = wb["GERAL"]
    
    # Limpa as abas antes de adicionar novos dados
    ws_dados.delete_rows(1, ws_dados.max_row)
    ws_geral.delete_rows(1, ws_geral.max_row)
    ws_geral.cell(row=1, column=1, value="AP")
    ws_geral.cell(row=1, column=2, value="ITEM")
    ws_geral.cell(row=1, column=3, value="TOTAIS")
    
    i = 1
    j = 2
    
    item_counts = defaultdict(int)
    
    for gstar_text in gstar_doc.ModelSpace:
        if gstar_text.ObjectName in ["AcDbText", "AcDbMText"]:
            if gstar_text.TextString.startswith(("(1)", "(2)", "(3)")):
                ws_dados.cell(row=i, column=1, value=gstar_text.TextString)
                i += 1
                
                prefixo = gstar_text.TextString
                numero = gstar_text.TextString[1:2]
                partes = gstar_text.TextString[3:].split()
                
                for parte in partes:
                    item_counts[(numero, parte)] += 1
    
    for (numero, parte), count in item_counts.items():
        ws_geral.cell(row=j, column=1, value=numero)
        ws_geral.cell(row=j, column=2, value=parte)
        ws_geral.cell(row=j, column=3, value=count)
        j += 1
    
    wb.save(excel_dados)
    wb.close()
    
    gstar_doc.Close(False)
    gstar_app.Quit()
#==================================================================================================================================================================================================================================================================================================================================================
def download_services_pdf(): #Baixar PDF das ordens de serviço - Comando_4
  # Verificar se o arquivo de dados de login existe substituido pela chamada "login, senha = get_login_data()"
  '''arquivo_login = "dados_login.txt"

  if not os.path.exists(arquivo_login):
    # Caso o arquivo não exista, solicitar o login e a senha via caixa de diálogo
    login = simpledialog.askstring("Typer", "Digite o BR:")
    senha = simpledialog.askstring("Typer", "Digite a senha:")

    # Gravar os dados de login no arquivo
    with open(arquivo_login, "w") as arquivo:
        arquivo.write(f"{login}\n")
        arquivo.write(f"{senha}\n")
    messagebox.showinfo("Sucesso!", "Dados gravados com sucesso!")
  else:
    # Ler os dados de login do arquivo
    with open(arquivo_login, "r") as arquivo:
      login = arquivo.readline().strip()
      senha = arquivo.readline().strip()'''


  #Leitura do arquivo ...

  # Inicializar a janela do Tkinter
  root = Tk()
  root.withdraw()  # Esconder a janela principal

  # Solicitar ao usuário para selecionar o arquivo Excel
  file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])

  if not file_path:
    print("Nenhum arquivo selecionado.")
  else:
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)

    # Selecionar a planilha desejada (substitua "Planilha1" pelo nome da sua planilha)
    sheet = workbook["Planilha1"]

    # Inicializar uma lista para armazenar os números
    numeros = []

    # Loop pelas células na coluna A a partir da segunda linha
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        sob = str(row[0]).zfill(10)  # Preencher com zeros à esquerda
        numeros.append(sob)

  # restante do codigo
  options = EdgeOptions()
  options.use_chromium = True
  driver = Edge(options=options)
  urllogin = 'http://gomnet-ce.enelint.global/'
  urlacomp_proj = 'http://gomnet-ce.enelint.global/EstudoTecnico.aspx'
  login, senha = get_login_data()
  #chrome_driver_path = 'C:\\Typer\\chromedriver.exe'
  #service = Service(chrome_driver_path)
  #driver = webdriver.Chrome(service=service)
  # Navega para a página de login
  driver.get(urllogin)
  # Localiza e preenche o campo de login
  campo_login = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'txtBoxLogin')))
  campo_login.send_keys(login)
  # Localiza e preenche o campo de senha
  campo_senha = driver.find_element(By.ID, 'txtBoxSenha') 
  campo_senha.send_keys(senha)
  # Localiza e clica no botão de login
  botao_login = driver.find_element(By.XPATH, '//*[@id="ImageButton_Login"]')  # Substitua pelo XPath correto
  botao_login.click()

  wait = WebDriverWait(driver, 10)
  elemento_alvo = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/form/table/tbody/tr[4]/td/div/div[1]/div/table/tbody/tr[5]/td/table/tbody/tr/td/input')))  # Substitua pelo ID correto

  driver.get (urlacomp_proj)

  # Limpar relatório
  wait = WebDriverWait(driver, 10)
  botao_limpar_relatorio = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_ImageButton_Limpar')))
  botao_limpar_relatorio.click()

  def baixar_sob(sob):    
    # Preenche o numero da sob
    elem = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_txtBoxNumSOB') 
    if elem.get_attribute('value'):
        elem.clear()
    elem.send_keys(sob)
    botao_enviar = driver.find_element(By.ID, 'ctl00_ContentPlaceHolder1_ImageButton_Enviar') 
    botao_enviar.click()
    
    # Baixa o pdf
    wait = WebDriverWait(driver, 10)
    botao_pdf = wait.until(EC.presence_of_element_located((By.ID, 'ctl00_ContentPlaceHolder1_Gridview_GomNet1_ctl02_ImageButton_PDF')))
    botao_pdf.click()
    time.sleep(2)
    # Move o pdf para outra pasta renomeado com o numero da sob
    pasta_origem = r'C:\\Downloads'
    pasta_destino = r'C:\\Typer\\PDF_ORDENS'
    # Obtém uma lista de todos os arquivos na pasta de origem
    arquivos = os.listdir(pasta_origem)
    # Filtra apenas os arquivos (excluindo pastas)
    arquivos = [arquivo for arquivo in arquivos if os.path.isfile(os.path.join(pasta_origem, arquivo))]
    # Classifica os arquivos por data de modificação (do mais recente para o mais antigo)
    arquivos.sort(key=lambda x: os.path.getmtime(os.path.join(pasta_origem, x)), reverse=True)
    # Verifica se existem arquivos na pasta
    if arquivos:
        # Obtém o caminho completo do último arquivo baixado
        ultimo_arquivo = os.path.join(pasta_origem, arquivos[0])
        time.sleep(1)
        # Move o último arquivo baixado para a pasta de destino
        shutil.move(ultimo_arquivo, pasta_destino)
        time.sleep(1)
        caminho_arquivo = 'C:\\Typer\\PDF_ORDENS\\Orcamento_.pdf'
        novo_nome_arquivo = 'C:\\Typer\\PDF_ORDENS\\' + str(sob) + ' SOB.pdf'
        # Renomear o arquivo
        os.rename(caminho_arquivo, novo_nome_arquivo)
    else:
      print("Nenhum arquivo encontrado na pasta de origem.")

  #messagebox.showinfo("Atenção!", "Antes de prosseguir, certifique-se de que não haja pdf de sobs pré-existentes no diretório (C:\\Downloads)!")

  #Loop para baixar cada pdf da lista de sobs
  for sob in numeros:
    baixar_sob(sob)
    time.sleep (2)
  messagebox.showinfo("Sucesso!", "Ordens de serviço baixadas com sucesso!")

  driver.quit()
#==================================================================================================================================================================================================================================================================================================================================================
def gerar_valoração_na_gom(): # Comando_5
  # Caminho para o arquivo Excel
  arquivo_excel = r"C:\\Typer\\Controle e Orçamento Boslan 2024-1-0.xlsm"

  # Abrir o arquivo Excel e selecionar a aba "VAL_MODELO"
  wb = xw.Book(arquivo_excel)
  sheet = wb.sheets["VAL_MODELO"]

  # Coletar os valores das células especificadas
  sob = sheet.range("o32").value
  #sob = '  '  #ordem cancelada
  # Garantir que `sob` seja uma string com 10 caracteres
  if sob.isdigit():  # Caso `sob` seja apenas números
    sob = sob.zfill(10)  # Preenche com zeros à esquerda até ter 10 dígitos
  else:
    sob = sob.rjust(10, '0')  # Para valores alfanuméricos, preenche à esquerda até 10 caracteres
        
  nome_obra = sheet.range("E1").value
  #nome_obra = '0059629997 EXT MT SETOR MAO GROSSA 0 JARDIM'

  # Coletar os valores das células especificadas
  und_base = sheet.range("B12").value
  #sob = '0059635729'  #ordem cancelada

  # Garantir que o numero_solicitacao tenha 10 dígitos
  und_base = f"{int(und_base):007d}"
  #und_base = '0077110'


  url_inicial = 'http://gom-ce.enelint.global/loginweb/login.asp' #Tela de login
  nova_url = 'http://gom-ce.enelint.global/AyudaActiva/AA_Consulta_Solicitud_NNSS.asp' #Tela para informar a sob
  codigo_projeto = 'E202326142' #estudo da ordem
  url_valoração = f'http://gom-ce.enelint.global/Proyecto_ET/PR_Val_Pry_ET.asp?Retorno=p&Criterio=ProyCodigo*{codigo_projeto}&Accion=C'
  login, senha = get_login_data()

  #login
  driver = webdriver.Chrome()
  driver.get(url_inicial)
  driver.maximize_window()
  url_atual = driver.current_url
  print('A URL atual da página é:', url_atual)
  print('Realizando login')
  elem = driver.find_element(By.NAME,'Usuario')
  elem.clear()
  elem.send_keys(login)
  elem = driver.find_element(By.NAME,'Clave')
  elem.clear()
  elem.send_keys(senha)
  elem.send_keys(Keys.RETURN)
  print('Login realizado')
  time.sleep(1)
  #url_atual = driver.current_url
  #print('A URL atual da página é:', url_atual)

  # Acessando a pagina da valoração

  driver.get(url_valoração)

  print('Acessando valoração')

  # Localize o elemento pelo nome
  elemento = driver.find_element(By.NAME, 'b_nueva_val_uucc.gif')

  # Crie uma instância de ActionChains
  actions = ActionChains(driver)
  actions.move_to_element(elemento).perform()
  # Mova o cursor do mouse para o elemento e clique
  actions.move_to_element(elemento).click().perform()

  #limpando e reescrevedo o nome da obra
  elem = driver.find_element(By.NAME,'txtValDescripcion') 
  elem.clear()   #Limma o campo sob 
  elem.send_keys(nome_obra)


  #Preencher tipo de preço (PD)
  elem = driver.find_element(By.NAME,'txtTipoPrecioCodigo') 
  elem.send_keys('PD')

  #Preencher tipo de preço (PD)
  elem = driver.find_element(By.NAME,'txtUnbaCodigo') 
  elem.send_keys(und_base)


  #Preencher observação da obra
  elem = driver.find_element(By.NAME,'txtValObservacion') 
  elem.send_keys(nome_obra)

  #Preencher viagens (2)
  elem = driver.find_element(By.NAME,'txtNumViagens').send_keys('2')
  #elem.send_keys(Keys.RETURN)

  elem = driver.find_element(By.XPATH,'/html/body/form/table/tbody/tr[4]/td[2]/table[2]/tbody/tr[1]/td/a[1]')
  elem.click()
                                 
  time.sleep(2)

  elem = driver.find_element(By.NAME,'txtNumeroApartado0').send_keys('1')
  elem = driver.find_element(By.NAME,'txtNumeroApartado1').send_keys('1')
  elem = driver.find_element(By.NAME,'txtNumeroApartado2').send_keys('1')
  elem = driver.find_element(By.NAME,'txtNumeroApartado3').send_keys('1')
  elem = driver.find_element(By.NAME,'txtNumeroApartado4').send_keys('1')

  elem = driver.find_element(By.NAME,'txtTocoCodigo0').send_keys('A')
  elem = driver.find_element(By.NAME,'txtTocoCodigo1').send_keys('A')
  elem = driver.find_element(By.NAME,'txtTocoCodigo2').send_keys('A')
  elem = driver.find_element(By.NAME,'txtTocoCodigo3').send_keys('A')
  elem = driver.find_element(By.NAME,'txtTocoCodigo4').send_keys('A')

  elem = driver.find_element(By.NAME,'txtTactCodigo0').send_keys('XS')
  elem = driver.find_element(By.NAME,'txtTactCodigo1').send_keys('XS')
  elem = driver.find_element(By.NAME,'txtTactCodigo2').send_keys('XS')
  elem = driver.find_element(By.NAME,'txtTactCodigo3').send_keys('XS')
  elem = driver.find_element(By.NAME,'txtTactCodigo4').send_keys('XS')

  elem = driver.find_element(By.NAME,'txtUnbaContrato0').send_keys('JA10077110')
  elem = driver.find_element(By.NAME,'txtUnbaContrato1').send_keys('JA10077110')
  elem = driver.find_element(By.NAME,'txtUnbaContrato2').send_keys('JA10077110')
  elem = driver.find_element(By.NAME,'txtUnbaContrato3').send_keys('JA10077110')
  elem = driver.find_element(By.NAME,'txtUnbaContrato4').send_keys('JA10077110')

  # Click na seta para preencher próxima página dos apartados
  driver.execute_script('InformarLista("AVANCELISTA1", 0, 200, 5);')

  elem = driver.find_element(By.NAME,'txtNumeroApartado0').send_keys('1')
  elem = driver.find_element(By.NAME,'txtNumeroApartado1').send_keys('1')

  elem = driver.find_element(By.NAME,'txtTocoCodigo0').send_keys('A')
  elem = driver.find_element(By.NAME,'txtTocoCodigo1').send_keys('A')

  elem = driver.find_element(By.NAME,'txtTactCodigo0').send_keys('XS')
  elem = driver.find_element(By.NAME,'txtTactCodigo1').send_keys('XS')

  elem = driver.find_element(By.NAME,'txtUnbaContrato0').send_keys('JA10077110')
  elem = driver.find_element(By.NAME,'txtUnbaContrato1').send_keys('JA10077110')


  time.sleep(10)



  """#Alternando direto para a página da SOB
  print('Alterando para página da SOB')
  driver.get(nova_url)
  time.sleep(2)
  url_atual = driver.current_url
  print('A URL atual da página é:', url_atual)

  #limpando a matricula e preenchendo a sob
  elem = driver.find_element(By.NAME,'txtSlnsCodigo') 
  elem.clear()   #Limma o campo sob 
  elem.send_keys(sob)
  elem = driver.find_element(By.NAME,'txtPfisUser')
  elem.clear() # Limpar a matrícula
  elem.send_keys(Keys.RETURN)
  time.sleep(2)

  elem = driver.find_element(By.NAME,'optRadioListado')
  #elem.clear() # Limpar a matrícula
  elem.send_keys(Keys.RETURN)
  time.sleep(2)

  # Localize o elemento pelo nome
  elemento = driver.find_element(By.NAME, 'lnkEntidades')

  # Crie uma instância de ActionChains
  actions = ActionChains(driver)
  actions.move_to_element(elemento).perform()
  # Mova o cursor do mouse para o elemento e clique
  actions.move_to_element(elemento).click().perform()


  driver.get('http://gom-ce.enelint.global/Proyecto_ET/PR_Val_Pry_ET.asp?Retorno=p&Criterio=ProyCodigo*',estudo,'&Accion=C')

  time.sleep(1)

  print(' clicado')

  time.sleep(10)

  print(' codigo acabou ')

  time.sleep(10)"""
#==================================================================================================================================================================================================================================================================================================================================================
# Criar a janela principal
root = tk.Tk()
#root.state('zoomed')
root.title(f'{empresa} Consulta')
largura = 740
altura = 350
largura_tela = root.winfo_screenmmwidth()
altura_tela = root.winfo_screenmmheight()
pos_x = 810
pos_Y = 450
#pos_x = (largura_tela // 50 ) - (largura //2)
#pos_Y = (altura_tela // 20 ) - (altura //2)
#Definir as dimensões da janela (largura x altura + posição X + posição Y)
root.geometry(f'{largura}x{altura}+{pos_x}+{pos_Y}')

#font_style = ("Helvetica", 16, "bold")
label_ceco = tk.Label(root, text="CECO:")
label_ceco.place(x=15, y= 15)

var_ceco = tk.Label(root, text= '8531')
var_ceco.place(x=15, y= 35)



#entry_dpto = tk.Entry(root, width=8)
#entry_dpto.place(x=15, y=35)

label_dpto_desc = tk.Label(root, text= "METROPOLITANA") 
label_dpto_desc.place(x=60, y=35)

# Função para converter a entrada para maiúsculas
def to_uppercase(*args):
    entry_sob.set(entry_sob.get().upper())
    if sob.isdigit() and len(sob) < 10:  # Caso `sob` seja apenas números
      sob = sob.zfill(10)  # Preenche com zeros à esquerda até ter 10 dígitos
    else:
      sob = sob.rjust(10, '0')  # Para valores alfanuméricos, preenche à esquerda até 10 caracteres
    
    entry_sob.set(sob)

label_sob = tk.Label(root, text="SOB:")
label_sob.place(x=180, y=15)
entry_sob = tk.StringVar()
entry_sob.trace_add('write', to_uppercase)
entry = tk.Entry(root, width=15, textvariable=entry_sob)
entry.place(x=180, y=35)




'''label_datasol = tk.Label(root, text="Dt. Solicitação:")
label_datasol.place(x=300, y=15)
entry_datasol = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/MM/yyyy')
entry_datasol.place(x=300, y=35)

label_datadesp = tk.Label(root, text="Dt. Despacho:")
label_datadesp.place(x=400, y=15)
entry_datadesp = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/MM/yyyy')
entry_datadesp.place(x=400, y=35)

label_tempo = tk.Label(root, text="Tempo:")
label_tempo.place(x=500, y=15)
entry_datadesp = tk.Entry(root, width=6)
entry_datadesp.place(x=500, y=35)

label_datalim = tk.Label(root, text="Dt. Limite:")
label_datalim.place(x=600, y=15)
entry_datalim = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/MM/yyyy')
entry_datalim.place(x=600, y=35)'''

label_solicitante = tk.Label(root, text="Solicitante:")
label_solicitante.place(x=15, y=60)
entry_solicitante = tk.Entry(root,width=50)
entry_solicitante.place(x=15, y=80)

label_municipio = tk.Label(root, text="Município:")
label_municipio.place(x=330 , y=60 )
entry_municipio = tk.Entry(root, width= 30)
entry_municipio.place(x=330 , y=80 )

# Criar e posicionar o botão para buscar dados
button_fetch = tk.Button(root, text="CONSULTAR", command=fetch_data)
button_fetch.place(x=15 , y=100)

# Criar e posicionar o botão para buscar dados
button_fetch = tk.Button(root, text="Apagar Ordem", command=deletar_ordem_tkinter)
button_fetch.place(x=15 , y=127)

#Nona Janela
#==================================================================================================================================================================================================================================================================================================================================================
def functions():
  nova_janela = Toplevel(root)
  nova_janela.title(f'{empresa} - Functions')
  nova_janela.geometry("480x190")
  nova_janela.resizable(False, False)
    
  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Importar Vistoria", command=import_relat_vistoria)
  button_fetch.place(x= 15, y=15)

  # Botão alterar senha Gom / GomNet
  button_fetch = tk.Button(nova_janela, width=30, text="Alterar senha Gom / Gomnet", command=change_password_gom_gomnet)
  button_fetch.place(x= 245, y=15)

  # Botão importar dados do relatório de acompanhamento de projetos
  button_fetch = tk.Button(nova_janela, width=30, text="Importar 02 e 03", command=import_relat_02_e_03)
  button_fetch.place(x= 15, y=42)

  # Botão baixar relatório de acompanhamento de projetos - Comando_2
  button_fetch = tk.Button(nova_janela, width=30, text="Baixar Relatórios", command=download_reports_02_e_03)
  button_fetch.place(x= 245, y=42)

  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Verificar localidades", command=verificar_localidade)
  button_fetch.place(x= 15, y=69)

  # Botão extrair do dxf - Comando_3
  button_fetch = tk.Button(nova_janela, width=30, text="Extrair do DXF", command=extrair_do_dxf)
  button_fetch.place(x= 245, y=69)

  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Atualizar localidades via excel", command=atualizar_localidades)
  button_fetch.place(x= 15, y=96)

  # Botão baixar pdf das ordens - Comando_4
  button_fetch = tk.Button(nova_janela, width=30, text="Baixar PDF das Ordens", command=download_services_pdf)
  button_fetch.place(x= 245, y=96)

  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Atualizar localidades via gom", command=atualizar_localidades_gom)
  button_fetch.place(x= 15, y=123)

  # Botão gerar valoração na gom - Comando_5
  button_fetch = tk.Button(nova_janela, width=30, text="Gerar valoração na gom", command=gerar_valoração_na_gom)
  button_fetch.place(x= 245, y=123)

  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Excluir tabela processo", command=excluir_tabela_processo)
  button_fetch.place(x= 15, y=150)

  # Botão importar dados do relatório de vistoria
  button_fetch = tk.Button(nova_janela, width=30, text="Exportar tabela processo", command=exportar_tabela_processo)
  button_fetch.place(x= 245, y=150)
#==================================================================================================================================================================================================================================================================================================================================================
#Abre a nova janela
button_fetch = tk.Button(root, width=30, text="Functions", command=functions)
button_fetch.place(x= 240, y=312)

#functions()
root.mainloop()
#==================================================================================================================================================================================================================================================================================================================================================